"""
Скрипт для парсинга HTML-отчета PostgreSQL и сохранения данных в Excel
"""

import json
import re
import pandas as pd
from bs4 import BeautifulSoup
from pathlib import Path
import argparse
import glob
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# Глобальные настройки
DEFAULT_HTML_FILE = "20 RPS.html"  # Файл по умолчанию

# Настройки форматирования Excel
HEADER_COLOR_RGB = (144, 238, 144)  # Цвет заголовков (светло-зеленый)
ROW_COLOR_RGB = (240, 255, 240)     # Цвет строк данных (аквамарин)
ENABLE_AUTOFILTER = True            # Включить автофильтр
ENABLE_AUTOFIT_COLUMNS = True       # Автоматически растягивать столбцы


def extract_data_from_html(html_file_path):
    """
    Извлекает JSON данные из HTML файла
    
    Args:
        html_file_path: путь к HTML файлу
        
    Returns:
        dict: извлеченные данные
    """
    print(f"Читаю файл: {html_file_path}")
    
    with open(html_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Ищем начало данных: const data=
    start_marker = 'const data='
    start_idx = content.find(start_marker)
    
    if start_idx == -1:
        raise ValueError("Не удалось найти 'const data=' в HTML файле")
    
    # Начинаем парсинг с открывающей скобки
    start_idx += len(start_marker)
    
    # Ищем конец JSON объекта с учетом вложенности скобок
    brace_count = 0
    in_string = False
    escape_next = False
    end_idx = start_idx
    
    for i in range(start_idx, len(content)):
        char = content[i]
        
        if escape_next:
            escape_next = False
            continue
            
        if char == '\\':
            escape_next = True
            continue
            
        if char == '"' and not escape_next:
            in_string = not in_string
            continue
            
        if not in_string:
            if char == '{':
                brace_count += 1
            elif char == '}':
                brace_count -= 1
                if brace_count == 0:
                    end_idx = i + 1
                    break
    
    data_str = content[start_idx:end_idx]
    data = json.loads(data_str)
    
    print(f"Найдено {len(data.get('datasets', {}))} наборов данных")
    
    return data


def parse_tables_from_html(html_file_path):
    """
    Парсит таблицы из HTML с помощью BeautifulSoup
    
    Args:
        html_file_path: путь к HTML файлу
        
    Returns:
        list: список DataFrame с таблицами
    """
    print(f"Парсинг HTML таблиц из {html_file_path}")
    
    with open(html_file_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
    
    # Находим все таблицы
    tables = soup.find_all('table')
    print(f"Найдено таблиц в HTML: {len(tables)}")
    
    return tables


def apply_sheet_formatting(worksheet, num_rows, num_cols):
    """
    Применяет форматирование к листу Excel
    
    Args:
        worksheet: лист Excel (openpyxl worksheet)
        num_rows: количество строк данных
        num_cols: количество столбцов
    """
    # Создаем заливки для цветов
    header_fill = PatternFill(
        start_color='{:02X}{:02X}{:02X}'.format(*HEADER_COLOR_RGB),
        end_color='{:02X}{:02X}{:02X}'.format(*HEADER_COLOR_RGB),
        fill_type='solid'
    )
    
    row_fill = PatternFill(
        start_color='{:02X}{:02X}{:02X}'.format(*ROW_COLOR_RGB),
        end_color='{:02X}{:02X}{:02X}'.format(*ROW_COLOR_RGB),
        fill_type='solid'
    )
    
    # Применяем форматирование к заголовкам (первая строка)
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
    
    # Применяем форматирование к строкам данных
    for row in range(2, num_rows + 2):  # +2 потому что заголовок в строке 1, данные начинаются со строки 2
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = row_fill
    
    # Включаем автофильтр
    if ENABLE_AUTOFILTER and num_rows > 0:
        worksheet.auto_filter.ref = f'A1:{get_column_letter(num_cols)}{num_rows + 1}'
    
    # Автоматическое растягивание столбцов
    if ENABLE_AUTOFIT_COLUMNS:
        for col in range(1, num_cols + 1):
            column_letter = get_column_letter(col)
            
            # Находим максимальную ширину в столбце
            max_length = 0
            for row in range(1, num_rows + 2):
                cell = worksheet.cell(row=row, column=col)
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Устанавливаем ширину столбца с небольшим запасом
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            worksheet.column_dimensions[column_letter].width = adjusted_width


def save_to_excel(data, output_file):
    """
    Сохраняет данные в Excel файл
    
    Args:
        data: словарь с данными
        output_file: путь к выходному Excel файлу
    """
    print(f"Сохранение данных в Excel: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Сохраняем основные свойства
        if 'properties' in data:
            props_df = pd.DataFrame([data['properties']])
            props_df.to_excel(writer, sheet_name='Properties', index=False)
            
            # Применяем форматирование
            worksheet = writer.sheets['Properties']
            apply_sheet_formatting(worksheet, len(props_df), len(props_df.columns))
            
            print(f"  ✓ Лист 'Properties' создан")
        
        # Сохраняем каждый dataset
        if 'datasets' in data:
            for dataset_name, dataset_values in data['datasets'].items():
                if isinstance(dataset_values, list) and len(dataset_values) > 0:
                    try:
                        # Обработка вложенных данных
                        df = pd.json_normalize(dataset_values)
                        
                        # Ограничение длины имени листа (Excel максимум 31 символ)
                        sheet_name = dataset_name[:31]
                        
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Применяем форматирование
                        worksheet = writer.sheets[sheet_name]
                        apply_sheet_formatting(worksheet, len(df), len(df.columns))
                        
                        print(f"  ✓ Лист '{sheet_name}' создан ({len(df)} строк, {len(df.columns)} столбцов)")
                    except Exception as e:
                        print(f"  ✗ Ошибка при обработке '{dataset_name}': {e}")
        
        # Сохраняем информацию о секциях
        if 'sections' in data:
            try:
                sections_df = pd.json_normalize(data['sections'])
                sections_df.to_excel(writer, sheet_name='Sections', index=False)
                
                # Применяем форматирование
                worksheet = writer.sheets['Sections']
                apply_sheet_formatting(worksheet, len(sections_df), len(sections_df.columns))
                
                print(f"  ✓ Лист 'Sections' создан")
            except Exception as e:
                print(f"  ✗ Ошибка при обработке секций: {e}")
    
    print(f"\n✓ Данные успешно сохранены в {output_file}")


def process_html_file(html_file_path):
    """Обрабатывает один HTML файл"""
    html_file = Path(html_file_path)
    
    if not html_file.exists():
        print(f"❌ Ошибка: файл {html_file} не найден!")
        return False
    
    # Выходной Excel файл
    output_file = html_file.with_suffix('.xlsx')
    
    print("=" * 70)
    print(f"Обработка файла: {html_file.name}")
    print("=" * 70)
    print()
    
    try:
        # Извлекаем данные
        data = extract_data_from_html(html_file)
        
        # Сохраняем в Excel
        save_to_excel(data, output_file)
        
        print()
        print("=" * 70)
        print(f"✓ Файл {html_file.name} успешно обработан!")
        print(f"  Создан: {output_file.name}")
        print("=" * 70)
        print()
        
        return True
        
    except Exception as e:
        print(f"\n❌ Ошибка при обработке {html_file.name}: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Основная функция"""
    # Парсинг аргументов командной строки
    parser = argparse.ArgumentParser(
        description='Парсинг HTML-отчетов PostgreSQL в Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры использования:
  %(prog)s                           # Использовать файл по умолчанию
  %(prog)s report.html               # Обработать один файл
  %(prog)s *.html                    # Обработать все HTML файлы в текущей папке
  %(prog)s "20 RPS.html" "40 RPS.html"  # Обработать несколько конкретных файлов
  %(prog)s C:/reports/*.html         # Обработать файлы по пути с маской
        """
    )
    
    parser.add_argument(
        'files',
        nargs='*',
        help=f'Путь к HTML файлу(ам) или маска (*.html). По умолчанию: {DEFAULT_HTML_FILE}'
    )
    
    args = parser.parse_args()
    
    # Определяем список файлов для обработки
    files_to_process = []
    
    if args.files:
        # Обрабатываем каждый аргумент
        for file_pattern in args.files:
            # Проверяем, содержит ли паттерн wildcards
            if '*' in file_pattern or '?' in file_pattern:
                # Используем glob для поиска файлов
                matched_files = glob.glob(file_pattern)
                if matched_files:
                    files_to_process.extend(matched_files)
                else:
                    print(f"⚠️ Предупреждение: паттерн '{file_pattern}' не совпал ни с одним файлом")
            else:
                # Обычный файл
                files_to_process.append(file_pattern)
    else:
        # Используем файл по умолчанию
        default_path = Path(__file__).parent / DEFAULT_HTML_FILE
        files_to_process.append(str(default_path))
    
    if not files_to_process:
        print("❌ Ошибка: не указаны файлы для обработки!")
        parser.print_help()
        return
    
    print()
    print("=" * 70)
    print("Парсинг HTML отчетов PostgreSQL в Excel")
    print("=" * 70)
    print(f"\nНайдено файлов для обработки: {len(files_to_process)}\n")
    
    # Обрабатываем каждый файл
    success_count = 0
    failed_count = 0
    
    for html_file in files_to_process:
        if process_html_file(html_file):
            success_count += 1
        else:
            failed_count += 1
    
    # Итоговая статистика
    print()
    print("=" * 70)
    print("Итоги обработки:")
    print("=" * 70)
    print(f"✓ Успешно обработано: {success_count}")
    if failed_count > 0:
        print(f"✗ Ошибок: {failed_count}")
    print("=" * 70)


if __name__ == "__main__":
    main()
